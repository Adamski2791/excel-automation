import os 
import openpyxl 

workbook = openpyxl.load_workbook(r"data/input/Healthcare_Appointments_Dataset.xlsx")
sheet = workbook['Healthcare_Appointments_Dataset']

print(sheet.title)



# delete row with missing value in the column 'PatientId'
def delete_missing_patient_id_rows(sheet):
    for row in sheet.iter_rows():
        if row[1].value == None:
            sheet.delete_rows(row[0].row)

def apply_currency_format(sheet):
    for row in sheet.iter_rows():
        if row[8].value is not None:
            # Set the number format to Euro currency
            row[8].number_format = '€#,##0.00'
            # Set the value to the number format
            row[8].value = row[8].value

# Add an additional column that wether the patient has an insurance or not
def add_column(sheet, column_name):
    sheet.insert_cols(11)
    sheet['K1'] = column_name
    for row in sheet.iter_rows(min_row=2):
        if row[9].value == 'None':
            row[10].value = 'No'
        else:
            row[10].value = 'Yes'


# IF a follow up is scheduled, highlight the row in green
def highlight_follow_ups(sheet):
    for row in sheet.iter_rows(min_row=2):
        if row[11].value == True:
            for cell in row:
                cell.fill = openpyxl.styles.PatternFill(
                    start_color='FF77DD77', 
                    end_color='FF77DD77', 
                    fill_type='solid')
                
# Apply excel calculation to calculate the consulatation fee after tax 
def apply_tax(sheet):
    for row in sheet.iter_rows(min_row=2):
        if row[8].value != None:
            row[8].value = float(row[8].value) * 1.15


# Save a diferent work book for each insurance company 
def save_workbook_per_insurance(sheet):
    insurance_companies = sheet['J']
    insurance_companies = set([cell.value for cell in insurance_companies])
    for company in insurance_companies:
        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.active
        new_sheet.title = company
        #add title row to new sheet with company nam
        
        new_sheet.append([cell.value for cell in sheet[1]]) # Copy the header
        for row in sheet.iter_rows(min_row=2):
            if row[9].value == company:
                new_sheet.append([cell.value for cell in row])
        new_workbook.save(f'data/transformed/{company}.xlsx')

def add_summary(workbook, sheet):
    diagnosis_fees = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):

        diagnosis = str(row[4])
        consultation_fee = float(row[8])

        if diagnosis in diagnosis_fees:
            diagnosis_fees[diagnosis] += consultation_fee
        else:
            diagnosis_fees[diagnosis] = consultation_fee

    summary_sheet = workbook.create_sheet('Summary')
    summary_sheet.append(['Diagnosis', 'Total Fee'])

    for diagnosis, fee in diagnosis_fees.items():
        summary_sheet.append([diagnosis, fee])


def add_chart(workbook):
    summary_sheet = workbook['Summary']
    first_sheet = workbook[workbook.sheetnames[0]]
    chart = openpyxl.chart.BarChart()
    chart.title = 'Diagnosis Fees'
    chart.x_axis.title = 'Diagnosis'
    chart.y_axis.title = 'Total Fee'

    labels = openpyxl.chart.Reference(summary_sheet, min_col=1, min_row=2, max_row=summary_sheet.max_row)
    data = openpyxl.chart.Reference(summary_sheet, min_col=2, min_row=1, max_row=summary_sheet.max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels) # be sure to specify this after add_data

    first_sheet.add_chart(chart, 'M1')


def run_all(input_path, output_path, new_column_name, sheet_name=None):
    for file in os.listdir(input_path):
        if file.endswith('.xlsx'):
            file_path = os.path.join(input_path, file)
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook[sheet_name] if sheet_name else workbook.active

            # Apply all processing functions
            delete_missing_patient_id_rows(sheet)
            apply_currency_format(sheet)
            add_column(sheet, new_column_name)
            highlight_follow_ups(sheet)
            apply_tax(sheet)
            add_summary(workbook, sheet)
            add_chart(workbook)

            # Save the modified workbook
            output_file_path = os.path.join(output_path, file)
            workbook.save(output_file_path)
            print(f"Processed and saved: {output_file_path}")

run_all('data/input', 'data/transformed', 'Has Insurance')
